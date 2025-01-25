import asyncio
import os
import sys
import subprocess
from datetime import datetime

# Ensure playwright browsers are installed
def install_browsers():
    try:
        subprocess.check_call([sys.executable, '-m', 'playwright', 'install', '--with-deps'])
    except Exception as e:
        print(f"Browser installation error: {e}")
        sys.exit(1)

# Install browsers before importing
install_browsers()

from playwright.async_api import async_playwright
from openpyxl import Workbook, load_workbook

async def scrape_cruise_listings():
    async with async_playwright() as p:
        print("Launching browser...")
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page()
        
        print("Navigating to the page...")
        await page.goto('https://www.costacruises.co.uk/cruises.html?page=1#{!tag=destinationTag}destinationIds=PE&occupancy_GBP_anonymous=A&guestAges=30&guestBirthdates=1995-01-25&group.sort=departDate%20asc')
        
        print("Waiting for cruise tiles...")
        await page.wait_for_selector('.costa-itinerary-tile', timeout=30000)
        
        cruise_tiles = await page.query_selector_all('.costa-itinerary-tile')
        print(f"Found {len(cruise_tiles)} cruise tiles.")
        
        cruise_data = []
        for tile in cruise_tiles:
            try:
                title = await tile.query_selector('.costa-itinerary-tile__title')
                title_text = await title.inner_text() if title else 'N/A'
                
                ship = await tile.query_selector('.costa-itinerary-tile__ship')
                ship_text = await ship.inner_text() if ship else 'N/A'
                
                price_element = await tile.query_selector('.currency-GBP')
                price_text = await price_element.inner_text() if price_element else 'N/A'
                
                dates_element = await tile.query_selector('.costa-itinerary-tile__dates')
                dates_text = await dates_element.inner_text() if dates_element else 'N/A'
                
                duration_element = await tile.query_selector('.costa-itinerary-tile__days')
                duration_text = await duration_element.inner_text() if duration_element else 'N/A'
                
                cruise_data.append({
                    'title': title_text,
                    'ship': ship_text,
                    'price': price_text,
                    'dates': dates_text,
                    'duration': duration_text
                })
            except Exception as e:
                print(f"Error processing tile: {e}")
        
        print(f"Scraped {len(cruise_data)} cruises.")
        await browser.close()
        return cruise_data

def save_to_excel(data, filename="cruise_data.xlsx"):
    sheet_name = datetime.now().strftime("%Y-%m-%d")
    
    try:
        if os.path.exists(filename):
            workbook = load_workbook(filename)
        else:
            workbook = Workbook()
            # Remove default sheet
            default_sheet = workbook.active
            workbook.remove(default_sheet)
        
        sheet = workbook.create_sheet(sheet_name)
        
        headers = ['Title', 'Ship', 'Price', 'Dates', 'Duration']
        sheet.append(headers)
        
        for cruise in data:
            sheet.append([
                cruise['title'], 
                cruise['ship'], 
                cruise['price'], 
                cruise['dates'], 
                cruise['duration']
            ])
        
        workbook.save(filename)
        print(f"Data saved to {filename}.")
    except Exception as e:
        print(f"Error saving to Excel: {e}")

async def main():
    cruises = await scrape_cruise_listings()
    if cruises:
        save_to_excel(cruises)
    else:
        print("No cruise data found.")

if __name__ == '__main__':
    asyncio.run(main())
